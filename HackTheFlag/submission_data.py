import openpyxl 
wb = openpyxl.Workbook() 
sheet = wb.active  
c1 = sheet.cell(row = 1, column = 1) 

c1.value = "ANKIT"

c2 = sheet.cell(row= 1 , column = 2) 
c2.value = "RAI"

c3 = sheet['A2'] 
c3.value = "RAH"

c4 = sheet['B2'] 
c4.value = "RAI"

c4 = sheet['B5'] 
c4.value = "RAI"

wb.save("demo.xlsx") 

